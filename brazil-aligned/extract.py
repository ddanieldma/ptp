from pathlib import Path
from openpyxl import load_workbook  
import openpyxl

import pandas as pd
from collections.abc import Sequence

DATA_DIR = Path("data")
RAW_DIR = DATA_DIR / "raw"
INTERIM_DIR = DATA_DIR / "interim"

# Constants and setup
FILE_PATH = RAW_DIR / 'Brazil-Aligned and Non-Aligned All Presidents.xlsx'
SHEET_NAME = 'Cabinet & Bureaucracy'
AGENCY_COLUMN = 'D'

PARQUET_PATH = RAW_DIR / 'data-editada-parquet-raw.parquet'

def get_ws() -> openpyxl.worksheet.Worksheet:
    """
    Returns the worksheet for testing and function use.

    Returns
    -------
    openpyxl.worksheet.Worksheet
        The worksheet object of the `SHEET_NAME` of the `FILE_PATH` file.
    """
    wb = load_workbook(FILE_PATH)
    return wb[SHEET_NAME]

def get_df_from_excel(file_path: str | Path = None, sheet_name: str = None) -> pd.DataFrame:
    """
    Gets the dataframe from the excel file.

    Parameters
    ----------
    file_path: str | Path, optional
        The path to the excel file . If None, uses the default.
    
    sheet_name: str
        The name of the sheet with the dataframe. If none, uses default.
    """

    if not file_path:
        file_path = FILE_PATH
    if not sheet_name:
        sheet_name = SHEET_NAME

    return pd.read_excel(file_path, sheet_name)

def exists_parquet(parquet_path: Path | str) -> bool:
    """
    Says if the parquet file exists or not.

    Parameters
    ----------
    parquet_path: Path | str
        The path to the parquet file.

    Returns
    -------
    bool
        A bool saying if the file exists or not.
    """
    
    return Path(parquet_path).exists()

def clean_dataset(
    df: pd.DataFrame,
    categorized_rows_list: Sequence,
    columns: list[str] = None
    ) -> pd.DataFrame:
    """
    Transforms the dataset by selecting columns, filling missing values, 
    adding categories, and renaming columns.

    Parameters
    ----------
    df : pd.DataFrame
        The raw dataframe to be transformed

    categorized_rows_list : Sequence
        Sequence of rows that are categorized
    
    columns: list[str], optional
        List of columns to be selected. If None, get all the columns.
    """
    
    # Removing unnecessary columns
    if columns:
        try:
            df = df[columns].copy()
        except KeyError:
            raise(f"Column not found in dataframe!")

    # Propagating last valid observation in the columns (this will correctly
    # fill the President and Year columns)
    df = df.ffill()

    # Adding column with category based in the color
    categorized_rows_list = categorized_rows_list[1:] # The first row is the header of the df
    df["category"] = pd.Series(categorized_rows_list, index=df.index)

    # Setting year column to int
    df["Year"] = df["Year"].astype(int)

    # Renaming columns
    df = df.rename(columns={"% Concedico e Parcialmente": "conc_parc"})
    df.columns = df.columns.str.lower()

    return df

def get_color_index(cell: openpyxl.cell) -> str:
    """
    Returns the color index of the cell.
    """
    return cell.fill.start_color.index if cell.fill.start_color else None

def categorize_rows(ws: openpyxl.worksheet, column_letter: str, colors_dict: dict) -> list:
    """
    Categorizes rows based on cell colors in the specified column.
    """

    categories = []
    
    for row in ws.iter_rows():
        cell = row[ord(column_letter.upper()) - ord('A')] # Gets the column number using ord
        color = get_color_index(cell) 
        
        if color == colors_dict["amarelo"]:
            category = "contra"
        elif color == colors_dict["vermelho"]:
            category = "alinhada"
        else:
            category = "neutra"
        
        categories.append(category)
    
    return categories

def clean_dataset_from_excel(
        ws: openpyxl.worksheet.Worksheet,
        agency_column: str = 'D',
        columns: list[str] = None
    ) -> pd.DataFrame:
    """
    Get's the data set from excel and cleans it.

    Parameters
    ----------
    ws : Worksheet
        The worksheet with the data.

    agency_column : str
        The agency column. Default is 'D'

    columns: list[str], optional
        List of columns to be selected. If None, get all the columns.
    """

    colors_dict = {
        "branco": get_color_index(ws["D1"]),
        "amarelo": get_color_index(ws["D16"]),
        "vermelho": get_color_index(ws["D257"]),
    }

    categorized_rows_list = categorize_rows(ws, agency_column, colors_dict)

    df = get_df_from_excel()

    return clean_dataset(df, categorized_rows_list, columns)

def get_cleaned_data_from_excel():
    """
    Get the whole data directly from excel, cleaned.

    Returns
    -------
    pd.Dataframe
        The dataframe with the data.
    """
    
    return clean_dataset_from_excel(get_ws(), AGENCY_COLUMN)

def get_data() -> pd.DataFrame:
    """
    Gets the data, from either parquet or excel.
    """
    if exists_parquet(PARQUET_PATH):
        df = pd.read_parquet(PARQUET_PATH)
    else:
        return clean_dataset_from_excel()

    return df